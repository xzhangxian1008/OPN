# 原先的版本是将id直接拼接送进网络优化用户/物品特征
# 该版本先将id转换为embedding，用embedding拼接，然后优化用户/物品的特征

# %matplotlib inline
import os, time, pickle, argparse, sys
import pickle as pk
import pandas as pd
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import numpy as np
from openpyxl import load_workbook
from scipy.stats import beta
torch.set_printoptions(threshold=10000)
np.set_printoptions(threshold=np.inf)

parser = argparse.ArgumentParser(description='RSAutoML')
parser.add_argument('--Train_Method', type=str, default='AutoML', help='options: AutoML, Supervised')
parser.add_argument('--Policy_Type', type=int, default=1, help='options: 0, 1, 2, 3, 4, 5')
parser.add_argument('--Val_Type', type=str, default='last_batch', help='options: last_batch, last_random')
parser.add_argument('--Loss_Type', type=str, default='MSE_sigmoid', help='options: MSE_sigmoid   MSE_no_sigmoid  BCEWithLogitsLoss   CrossEntropyLoss')
parser.add_argument('--Data_Set', type=str, default='ml-latest', help='options: ml-20m ml-latest')
parser.add_argument('--Dy_Emb_Num', type=int, default=2, help='options: 1, 2')
parser.add_argument('--Reward_Base', type=str, default=None, help='options: None, last_loss, ave_loss')
parser.add_argument('--last_num', type=int, default=5, help='options: 1, 2')
args = parser.parse_args()

Model_Gpu  = torch.cuda.is_available()
device     = torch.device('cuda:0' if Model_Gpu else 'cpu')
DATA_PATH  = '../data'
DATA_SET   = args.Data_Set
Batch_Size = 500     # batch size
LR_model   = 0.003   # learning rate, for RS model
LR_darts   = 0.0001  # learning rate, for policy network
Epoch      = 1       # train epoch
Beta_Beta  = 20      # beta for Beta distribution
H_alpha    = 0       # for nn.KLDivLoss 0.001

# TODO 改参数 bin: MSE_sigmoid  multi: CrossEntropyLoss
if DATA_SET == 'ml-20m':
    Train_Size   = 15000000      # training dataset size
elif DATA_SET == 'ml-latest':
    Train_Size = 22000000  # training dataset size
Test_Size    = 5000000       # training dataset size
Emb_Size     = [2, 4, 8, 16, 64, 128]  # 1,2,4,8,16,32,64,128,256,512
cross_emb_idx = 2
cross_emb_size = Emb_Size[cross_emb_idx]  # for object-aware
Emb_Con_Size = [2+cross_emb_size, 4+cross_emb_size, 8+cross_emb_size, 16+cross_emb_size, 64+cross_emb_size, 128]
Train_Method = args.Train_Method
Policy_Type  = args.Policy_Type
Types        = ['Policy0: embedding for popularity',
                'Policy1: embedding for popularity + last_weights',
                'Policy2: embedding for popularity + last_weights + last_loss',
                'Policy3: popularity one_hot',
                'Policy4: popularity one_hot + last_weights',
                'Policy5: popularity one_hot + last_weights  + last_loss']
Val_Type     = args.Val_Type

Dy_Emb_Num   = args.Dy_Emb_Num       # dynamic num of embedding to adjust, 1 for user, 2 for user & movie
Reward_Base  = args.Reward_Base
last_num = args.last_num
Loss_Type    = args.Loss_Type
ControllerLoss = nn.CrossEntropyLoss(reduce=False)


print('\n****************************************************************************************\n')
print('os.getpid():   ', os.getpid())
if torch.cuda.is_available():
    print('torch.cuda:    ', torch.cuda.is_available(), torch.cuda.current_device(), torch.cuda.device_count(), torch.cuda.get_device_name(0), torch.cuda.device(torch.cuda.current_device()))
else:
    print('GPU is not available!!!')
print('Train_Size:    ', Train_Size)
print('Test_Size:     ', Test_Size)
print('Emb_Size:      ', Emb_Size)
print('Batch_Size:    ', Batch_Size)
print('Dy_Emb_Num:    ', Dy_Emb_Num)
print('Loss_Type:     ', Loss_Type)
print('Train_Method:  ', Train_Method)
print('Policy_Type:   ', Types[Policy_Type])
print('Val_Type:      ', Val_Type)
print('Beta_Beta:     ', Beta_Beta)
print('H_alpha:       ', H_alpha)
print('LR_model:      ', LR_model)
print('LR_darts:      ', LR_darts)
print('\n****************************************************************************************\n')


def load_data():
    train_features, test_features, train_target, test_target \
        = pickle.load(open('{}/{}_TrainTest_{}_{}.data'.format(DATA_PATH, DATA_SET, Train_Size, Output_Dim), mode='rb'))
    print("len(train_features): ", len(train_features))
    print("len(test_features): ", len(test_features))
    test_features, test_target = test_features[:Test_Size], test_target[:Test_Size]
    genome_scores_dict = pickle.load(open('{}/{}_GenomeScoresDict_{}.data'.format(DATA_PATH, DATA_SET, Output_Dim), mode='rb'))
    train_feature_data = pd.DataFrame(train_features, columns=['userId', 'movieId', 'user_frequency', 'movie_frequency'])
    test_feature_data = pd.DataFrame(test_features, columns=['userId', 'movieId', 'user_frequency', 'movie_frequency'])
    User_Num = max(train_feature_data['userId'].max() + 1, test_feature_data['userId'].max() + 1)  # 138494
    Movie_Num = max(train_feature_data['movieId'].max() + 1, test_feature_data['movieId'].max() + 1)  # 131263
    max_user_popularity = max(train_feature_data['user_frequency'].max()+1, test_feature_data['user_frequency'].max()+1)
    max_movie_popularity = max(train_feature_data['movie_frequency'].max() + 1, test_feature_data['movie_frequency'].max() + 1)
    return train_features, test_features, train_target, test_target, genome_scores_dict, \
           train_feature_data, test_feature_data, len(train_features), len(test_features), \
           User_Num, Movie_Num, max_user_popularity, max_movie_popularity


def Batch_Losses(Loss_Type, prediction, target):
    if Loss_Type == 'MSE_sigmoid':
        return nn.MSELoss(reduction='none')(nn.Sigmoid()(prediction), target)
    elif Loss_Type == 'MSE_no_sigmoid':
        return nn.MSELoss(reduction='none')(prediction, target)
    elif Loss_Type == 'BCEWithLogitsLoss':
        return nn.BCEWithLogitsLoss(reduction='none')(prediction, target)
    elif Loss_Type == 'CrossEntropyLoss':
        return nn.CrossEntropyLoss(reduction='none')(prediction, target)
    else:
        print('No such Loss_Type.')


def Batch_Accuracies(Loss_Type, prediction, target):
    with torch.no_grad():
        if Loss_Type == 'MSE_sigmoid':
            predicted = 1 * (torch.sigmoid(prediction).data > 0.5)
        elif Loss_Type == 'MSE_no_sigmoid':
            predicted = 1 * (prediction > 0.5)
        elif Loss_Type == 'BCEWithLogitsLoss':
            predicted = 1 * (torch.sigmoid(prediction).data > 0.5)
        elif Loss_Type == 'CrossEntropyLoss':
            _, predicted = torch.max(prediction, 1)
        else:
            print('No such Loss_Type.')

        Batch_Accuracies = 1 * (predicted == target)
        Batch_Accuracies = list(Batch_Accuracies.detach().cpu().numpy())
        return Batch_Accuracies


def Beta(length, popularity, be=10):
    x = [i/length for i in range(length+1)]
    cdfs = [beta.cdf(x[i+1], popularity, be) - beta.cdf(x[i], popularity, be) for i in range(length)]
    return cdfs


class Policy(nn.Module):
    def __init__(self, Setting_Popularity, Setting_Weight, Policy_Type):
        super(Policy, self).__init__()
        self.Policy_Type = Policy_Type
        if self.Policy_Type == 0:
            self.transfrom_input_length = Setting_Popularity[1]
        elif self.Policy_Type == 1:
            self.transfrom_input_length = Setting_Popularity[1] + Setting_Weight[1]
        elif self.Policy_Type == 2:
            self.transfrom_input_length = Setting_Popularity[1] + Setting_Weight[1] + 1
        elif self.Policy_Type == 3:
            self.transfrom_input_length = Setting_Popularity[0]
        elif self.Policy_Type == 4:
            self.transfrom_input_length = Setting_Popularity[0] + Setting_Weight[1]
        elif self.Policy_Type == 5:
            self.transfrom_input_length = Setting_Popularity[0] + Setting_Weight[1] + 1
        else:
            print('No such Policy_Type 1')

        if self.Policy_Type in [0, 1, 2]:
            self.emb_popularity = nn.Embedding(num_embeddings=Setting_Popularity[0], embedding_dim=Setting_Popularity[1])
            self.batch_norm = nn.BatchNorm1d(Setting_Popularity[1])
        elif self.Policy_Type in [3, 4, 5]:
            self.emb_popularity = nn.Embedding(num_embeddings=Setting_Popularity[0], embedding_dim=Setting_Popularity[0]).to(dtype=torch.float32)
            self.emb_popularity.weight.data = torch.eye(Setting_Popularity[0])
            self.emb_popularity.weight.requires_grad = False
        else:
            print('No such Policy_Type 2')

        self.transfrom = nn.Sequential(
            nn.Linear(self.transfrom_input_length, 512),
            nn.BatchNorm1d(512),
            nn.Tanh(),
            nn.Linear(512, 2)) # remain dim, increase dim
            # nn.Softmax(dim=1))

    # popularity：出现频率
    # emb_sizes：上一次的embedding大小
    # last_loss：上一次的损失值
    def forward(self, popularity, emb_sizes, last_loss):
        '''
        popularity: (batch_size)
        emb_sizes: (batch_size)
        output: (batch_size x 3)
        '''

        # emb_popularity: (batch_size x emb_size)
        emb_popularity = self.emb_popularity(popularity)
        if self.Policy_Type in [0, 1, 2]:
            transformed_emb_popularity = self.batch_norm(emb_popularity)
        elif self.Policy_Type in [3, 4, 5]:
            transformed_emb_popularity = emb_popularity
        else:
            transformed_emb_popularity = None
            print('No such Policy_Type 3')

        if self.Policy_Type in [0, 3]:
            concatenation = transformed_emb_popularity
        elif self.Policy_Type in [1, 4]:
            last_weights = nn.functional.one_hot(emb_sizes, num_classes=len(Emb_Size)).float()
            concatenation = torch.cat((transformed_emb_popularity, last_weights), 1)
        elif self.Policy_Type in [2, 5]:
            last_weights = nn.functional.one_hot(emb_sizes, num_classes=len(Emb_Size)).float()
            concatenation = torch.cat((transformed_emb_popularity, last_weights, last_loss), 1)
        else:
            print('No such Policy_Type 4')
        return self.transfrom(concatenation) # 返回一个Embedding大小调整方案，扩大或不变

# 这个是Linear transform部分，负责把不同embedding size的embedding结果转换成相同的大小，然后对两个emb做连接
class RS_MLP(nn.Module):
    def __init__(self, Output_Dim, Dynamic_Emb_Num):
        super(RS_MLP, self).__init__()
        # self.emb_user = nn.Embedding(num_embeddings=User_Num, embedding_dim=sum(Emb_Size))
        # self.emb_movie = nn.Embedding(num_embeddings=Movie_Num, embedding_dim=sum(Emb_Size))
        self.emb_user = nn.ModuleList(nn.Embedding(num_embeddings=User_Num, embedding_dim=emb_size) for emb_size in Emb_Size)
        self.emb_movie = nn.ModuleList(nn.Embedding(num_embeddings=Movie_Num, embedding_dim=emb_size) for emb_size in Emb_Size)
        self.cross_emb_user = nn.Embedding(num_embeddings=User_Num, embedding_dim=cross_emb_size)
        self.cross_emb_movie = nn.Embedding(num_embeddings=Movie_Num, embedding_dim=cross_emb_size)
        # for emb in self.emb_user + self.emb_movie:
        #     emb.to(device)
        self.bn_user = nn.BatchNorm1d(max(Emb_Size))
        self.bn_movie = nn.BatchNorm1d(max(Emb_Size))

        # 长度最大的embedding不会被转换，所以我们要加上一个转换模块
        self.user_cross_linear = nn.Linear(Emb_Size[len(Emb_Size)-1] + cross_emb_size, Emb_Size[len(Emb_Size)-1])
        self.movie_cross_linear = nn.Linear(Emb_Size[len(Emb_Size)-1] + cross_emb_size, Emb_Size[len(Emb_Size)-1])

        # self.user_bn = nn.BatchNorm1d(Emb_Size[len(Emb_Size) - 1])
        # self.movie_bn = nn.BatchNorm1d(Emb_Size[len(Emb_Size) - 1])

        # ModuleList是一个列表，内部可以放很多模型，用的时候指定索引就可以了
        self.W_user = nn.ModuleList([nn.Linear(Emb_Con_Size[i], Emb_Con_Size[i + 1]) for i in range(len(Emb_Con_Size) - 1)])
        self.W_movie = nn.ModuleList([nn.Linear(Emb_Con_Size[i], Emb_Con_Size[i + 1]) for i in range(len(Emb_Con_Size) - 1)])

        self.my_W_user = nn.ModuleList([nn.Linear(Emb_Size[i], Emb_Size[i + 1]) for i in range(len(Emb_Size) - 1)])
        self.my_W_movie = nn.ModuleList([nn.Linear(Emb_Size[i], Emb_Size[i + 1]) for i in range(len(Emb_Size) - 1)])
        self.tanh = nn.Tanh()
        self.movie_transfrom = nn.Sequential(  # nn.BatchNorm1d(1128),
            nn.Linear(1128, 512),
            nn.BatchNorm1d(512),
            nn.Tanh(),
            nn.Linear(512, max(Emb_Size)))
        self.transfrom = nn.Sequential(
            nn.BatchNorm1d(max(Emb_Size) * 2),
            nn.Linear(max(Emb_Size) * 2, 512),
            nn.BatchNorm1d(512),
            nn.Tanh(),
            nn.Linear(512, Output_Dim))
        self.den = Dynamic_Emb_Num
        # setattr(self, 'z', 666)

    # 注意，这里的size都是索引值
    def forward(self, u_emb_sizes, m_emb_sizes, userId, movieId, movie_vec):
        '''
        u_emb_sizes: (batch_size)
        m_emb_sizes: (batch_size)
        '''

        # u_weight: (batch_size x emb_num)
        # m_weight: (batch_size x emb_num)
        # num_classes：指定one_hot的最大长度
        u_weight = nn.functional.one_hot(u_emb_sizes, num_classes=len(Emb_Size))
        if self.den == 2:
            m_weight = nn.functional.one_hot(m_emb_sizes, num_classes=len(Emb_Size))

        # user_emb是一个列表，存储了用户所有大小的embedding结果
        # movie_emb也是类似的
        user_emb = [self.emb_user[i](userId) for i in range(len(Emb_Size))]
        # cross_user_emb = self.cross_emb_user(userId)
        movie_emb = None if self.den == 1 else [self.emb_movie[i](movieId) for i in range(len(Emb_Size))]
        # cross_movie_emb = self.cross_emb_movie(movieId)

        user_embs = []
        for i in range(len(Emb_Size)):
            # temp = user_emb[i]
            temp = torch.cat((user_emb[i], self.emb_movie[cross_emb_idx](movieId)), 1)
            if i == len(Emb_Size) - 1:
                temp = self.user_cross_linear(temp)

            # 这层for循环就是把PN给出来的不规则大小的embedding经过MLP层层映射最后转换成相同大小的embedding
            for j in range(i, len(Emb_Size) - 1):
                temp = self.W_user[j](temp)
            user_embs.append(temp) # 最后的值加入到这里面

        if self.den == 2:
            movie_embs = []
            for i in range(len(Emb_Size)):
                # temp = movie_emb[i]
                temp = torch.cat((movie_emb[i], self.emb_user[cross_emb_idx](userId)), 1)
                if i == len(Emb_Size) - 1:
                    temp = self.movie_cross_linear(temp)

                for j in range(i, len(Emb_Size) - 1):
                    temp = self.W_movie[j](temp)
                movie_embs.append(temp)

        v_user = sum([torch.reshape(u_weight[:, i], (len(u_weight), -1)) * self.tanh(
            self.bn_user(user_embs[i])) for i in range(len(Emb_Size))])
        v_movie = sum([torch.reshape(m_weight[:, i], (len(m_weight), -1)) * self.tanh(
            self.bn_movie(movie_embs[i])) for i in range(len(Emb_Size))]) if self.den == 2 else self.movie_transfrom(movie_vec)

        user_movie = torch.cat((v_user, v_movie), 1)
        return self.transfrom(user_movie)


def update_controller(index, features, target):
    """ Update user_policy and movie_policy """
    if Train_Method == 'AutoML' and index > 0:
        if Val_Type == 'last_random':
            val_index = np.random.choice(index, Batch_Size)
            batch_train = features[:index][val_index]
            batch_train_target = target[:index][val_index]
        elif Val_Type == 'last_batch':
            batch_train = features[index - Batch_Size:index]
            batch_train_target = target[index - Batch_Size:index]
        else:
            batch_train = None
            batch_train_target = None
            print('No such Val_Type')

        # [:, 0]表示取第0列数据
        userId = torch.tensor(batch_train[:, 0], requires_grad=False).to(device)
        movieId = torch.tensor(batch_train[:, 1], requires_grad=False).to(device)
        userPop = torch.tensor(batch_train[:, 2], requires_grad=False).to(device)
        moviePop = torch.tensor(batch_train[:, 3], requires_grad=False).to(device)
        old_uw = torch.tensor(user_weights[batch_train[:, 0]], requires_grad=False).to(device)
        old_mw = torch.tensor(movie_weights[batch_train[:, 1]], requires_grad=False).to(device)
        old_ul = torch.tensor(user_losses[batch_train[:, 0], :], requires_grad=False).to(device)
        old_ml = torch.tensor(movie_losses[batch_train[:, 1], :], requires_grad=False).to(device)
        movie_vec = torch.tensor([genome_scores_dict[str(batch_train[:, 1][i])] for i in range(len(batch_train[:, 1]))],
                                 requires_grad=False).to(device) if Dy_Emb_Num == 1 else None
        batch_train_target = torch.tensor(batch_train_target,
                                          dtype=torch.int64 if Loss_Type == 'CrossEntropyLoss' else torch.float32,
                                          requires_grad=False).to(device)

        if Reward_Base == 'ave_loss':
            old_utl = torch.tensor(user_total_losses[batch_train[:, 0]], requires_grad=False).to(device)
            old_mtl = torch.tensor(movie_total_losses[batch_train[:, 1]], requires_grad=False).to(device)
            old_uc = torch.tensor(user_count[batch_train[:, 0]], requires_grad=False).to(device)
            old_mc = torch.tensor(movie_count[batch_train[:, 1]], requires_grad=False).to(device)

        # user_adj_prob: (batch_size x 2)
        # userPop：用户的频率
        # old_uw：老旧的用户embedding大小
        # user_adj_weights：模型给出的初步信息，还要经过softmax才能下结论
        user_adj_weights = user_policy(userPop, old_uw, old_ul)

        # 得出结论，是enlarge还是unchange
        # user_adj_prob：代表模型给出的建议
        user_adj_prob = nn.functional.softmax(user_adj_weights, dim=-1)

        # 要注意，这里的weight都是索引值
        mask = old_uw != len(Emb_Size) - 1

        # user_adj_samples: (batch_size)
        # user_adj_samples：embedding变化量
        user_adj_samples = mask * torch.multinomial(user_adj_prob, 1).squeeze(1)

        # new_uw: (batch_size)
        # new_uw：调整后用户的embedding大小
        new_uw = old_uw + user_adj_samples

        if Dy_Emb_Num == 2:
            # movie_adj_prob: (batch_size x 3)
            movie_adj_weights = movie_policy(moviePop, old_mw, old_ml)
            movie_adj_prob = nn.functional.softmax(movie_adj_weights, dim=-1)

            mask = old_mw != len(Emb_Size) - 1
            # movie_adj_samples: (batch_size)
            movie_adj_samples = mask * torch.multinomial(movie_adj_prob, 1).squeeze(1)

            # new_mw: (batch_size)
            new_mw = old_mw + movie_adj_samples

        else:
            new_mw = 0


        with torch.no_grad():
            temp_emb_user = model.emb_user

            for i in range(len(Emb_Size) - 1):
                j = i + 1
                part_userId = userId[
                    ((old_uw == i) * (new_uw == j)).nonzero().squeeze(1)]
                # 当发现emb需要被增加，模型会将原emb的参数扔进一层网络，得到的新参数被用在新维度的emb中
                # 也许作者认为这样可以保留先前的参数吧
                if len(part_userId) > 0:
                    # [part_userId, :]：取第part_userId行的数据
                    # Change  W_user -> my_W_user
                    model.emb_user[j].weight[part_userId, :] = model.my_W_user[i](
                        model.emb_user[i].weight[part_userId, :])

            if Dy_Emb_Num == 2:
                temp_emb_movie = model.emb_movie

                for i in range(len(Emb_Size) - 1):
                    j = i + 1
                    part_movieId = movieId[
                        ((old_mw == i) * (new_mw == j)).nonzero().squeeze(1)]
                    if len(part_movieId) > 0:
                        # Change  W_movie -> my_W_movie
                        model.emb_movie[j].weight[part_movieId, :] = model.my_W_movie[i](
                            model.emb_movie[i].weight[part_movieId, :])

            # rating: (batch_size x 1)
            # 在这里把Policy Network调整后的输出送进DRM模型，后面会根据DRM模型的结果来调整PN
            rating = model(new_uw, new_mw, userId, movieId, movie_vec)

            # 这下面都是在计算损失值了，为调整PN做准备

            # rating: (batch_size)
            rating = rating.squeeze(1).squeeze(1) if Loss_Type == 'CrossEntropyLoss' else rating.squeeze(1)

            # batch_losses: (batch_size)
            batch_losses = Batch_Losses(Loss_Type, rating, batch_train_target)
            rewards = 1 - batch_losses

        model.emb_user = temp_emb_user

        if Dy_Emb_Num == 1:
            if Reward_Base == 'last_loss':
                baseline = 1 - old_ul[:, 0]
                rewards = rewards - baseline
            elif Reward_Base == 'ave_loss':
                last_num_tensor = torch.Tensor([last_num]).repeat(len(old_uc)).to(device)
                baseline = 1 - torch.sum(old_utl, dim=1) / torch.where(old_uc < last_num, old_uc, last_num_tensor)
                rewards = rewards - baseline

            loss = torch.sum(ControllerLoss(user_adj_weights, user_adj_samples) * rewards)

            # if index % 100000 == 0:
            #     print("rewards: ", rewards[:50].tolist())

            optimizer_user.zero_grad()
            loss.backward()
            optimizer_user.step()

        elif Dy_Emb_Num == 2:
            model.emb_movie = temp_emb_movie
            if Reward_Base == 'last_loss':
                baseline_u = 1 - old_ul[:, 0]
                baseline_m = 1 - old_ml[:, 0]

            elif Reward_Base == 'ave_loss':
                last_num_tensor = torch.Tensor([last_num]).repeat(len(old_uc)).to(device)
                baseline_u = 1 - torch.sum(old_utl, dim=1) / torch.where(old_uc < last_num, old_uc, last_num_tensor)
                baseline_m = 1 - torch.sum(old_mtl, dim=1) / torch.where(old_mc < last_num, old_mc, last_num_tensor)

            rewards_u = rewards - baseline_u
            rewards_m = rewards - baseline_m

            loss_u = torch.sum(ControllerLoss(user_adj_weights, user_adj_samples) * rewards_u)
            loss_m = torch.sum(ControllerLoss(movie_adj_weights, movie_adj_samples) * rewards_m)

            # if index % 100000 == 0:
            #     print("rewards_u: ", rewards_u[:50].tolist())
            #     print("rewards_m: ", rewards_m[:50].tolist())

            optimizer_user.zero_grad()
            loss_u.backward()
            optimizer_user.step()
            optimizer_movie.zero_grad()
            loss_m.backward()
            optimizer_movie.step()


def update_RS(index, features, Len_Features, target, mode):
    """ Update RS's embeddings and NN """
    global train_sample_loss, train_sample_accuracy, user_dims_record, movie_dims_record
    index_end = index + Batch_Size
    if index_end >= Len_Features:
        batch_train = features[index:Len_Features]
        batch_train_target = target[index:Len_Features]
    else:
        batch_train = features[index:index_end]
        batch_train_target = target[index:index_end]

    userId = torch.tensor(batch_train[:, 0], requires_grad=False).to(device)
    movieId = torch.tensor(batch_train[:, 1], requires_grad=False).to(device)
    userPop = torch.tensor(batch_train[:, 2], requires_grad=False).to(device)
    moviePop = torch.tensor(batch_train[:, 3], requires_grad=False).to(device)
    old_uw = torch.tensor(user_weights[batch_train[:, 0]], requires_grad=False).to(device)
    old_mw = torch.tensor(movie_weights[batch_train[:, 1]], requires_grad=False).to(device)
    old_ul = torch.tensor(user_losses[batch_train[:, 0], :], requires_grad=False).to(device)
    old_ml = torch.tensor(user_losses[batch_train[:, 1], :], requires_grad=False).to(device)
    movie_vec = torch.tensor([genome_scores_dict[str(batch_train[:, 1][i])] for i in range(len(batch_train[:, 1]))],
                             requires_grad=False).to(device) if Dy_Emb_Num == 1 else None
    batch_train_target = torch.tensor(batch_train_target,
                                      dtype=torch.int64 if Loss_Type == 'CrossEntropyLoss' else torch.float32,
                                      requires_grad=False).to(device)

    with torch.no_grad():

        # user_adj_prob: (batch_size x 2)
        user_adj_weights = user_policy(userPop, old_uw, old_ul)
        user_adj_prob = nn.functional.softmax(user_adj_weights, dim=-1)

        mask = old_uw != len(Emb_Size) - 1

        user_adj = mask * torch.argmax(user_adj_prob, dim=1)

        # new_uw: (batch_size)
        new_uw = old_uw + user_adj

        # if index % 500000 == 0:
        #     print("old_uw: ", old_uw)
        #     print("new_uw: ", new_uw)


        for i in range(len(Emb_Size) - 1):
            j = i + 1

            part_userId = userId[((old_uw == i) * (new_uw == j)).nonzero().squeeze(1)]
            # 当发现emb需要被增加，模型会将原emb的参数扔进一层网络，得到的新参数被用在新维度的emb中
            # 也许作者认为这样可以保留先前的参数吧
            if len(part_userId) > 0:
                # Change  W_user -> my_W_user
                model.emb_user[j].weight[part_userId, :] = model.my_W_user[i](
                    model.emb_user[i].weight[part_userId, :])

        if Dy_Emb_Num == 2:
            # movie_adj_prob: (batch_size x 2)
            movie_adj_weights = movie_policy(moviePop, old_mw, old_ml)
            movie_adj_prob = nn.functional.softmax(movie_adj_weights, dim=-1)
            mask = old_mw != len(Emb_Size) - 1

            movie_adj = mask * torch.argmax(movie_adj_prob, dim=1)

            # new_mw: (batch_size x emb_num)
            new_mw = old_mw + movie_adj

            # if index % 500000 == 0:
            #     print("old_mw: ", old_mw)
            #     print("new_mw: ", new_mw)

            for i in range(len(Emb_Size) - 1):
                j = i + 1
                part_movieId = movieId[
                    ((old_mw == i) * (new_mw == j)).nonzero().squeeze(1)]
                if len(part_movieId) > 0:
                    # Change  W_movie -> my_W_movie
                    model.emb_movie[j].weight[part_movieId, :] = model.my_W_movie[i](
                        model.emb_movie[i].weight[part_movieId, :])
        else:
            new_mw = 0


    rating = model(new_uw, new_mw, userId, movieId, movie_vec)
    rating = rating.squeeze(1).squeeze(1) if Loss_Type == 'CrossEntropyLoss' else rating.squeeze(1)

    # batch_losses: (batch_size)
    batch_losses = Batch_Losses(Loss_Type, rating, batch_train_target)
    loss = sum(batch_losses)
    batch_accuracies = Batch_Accuracies(Loss_Type, rating, batch_train_target)

    train_sample_loss += list(batch_losses.detach().cpu().numpy())
    losses[mode].append(loss.detach().cpu().numpy())
    train_sample_accuracy += batch_accuracies
    accuracies[mode].append((sum(batch_accuracies), len(batch_train_target)))

    user_dims_record += [Emb_Size[item] for item in new_uw.detach().cpu()]
    if Dy_Emb_Num == 2:
        movie_dims_record += [Emb_Size[item] for item in new_mw.detach().cpu()]

    if Train_Method == 'AutoML':
        optimizer_model.zero_grad()
        loss.backward()
        optimizer_model.step()
    elif Train_Method == 'Supervised':
        optimizer_whole.zero_grad()
        loss.backward()
        optimizer_whole.step()
    else:
        print('No such Train_Method')


    """ Update old_uw old_mw old_ul old_ml """

    user_weights[batch_train[:, 0]] = new_uw.detach().cpu().numpy()
    movie_weights[batch_train[:, 1]] = new_mw.detach().cpu().numpy() if Dy_Emb_Num == 2 else np.zeros((len(batch_train),))
    user_losses[batch_train[:, 0], :] = np.reshape(batch_losses.detach().cpu().numpy(), (-1, 1))
    movie_losses[batch_train[:, 1], :] = np.reshape(batch_losses.detach().cpu().numpy(), (-1, 1))
    final_user_pop[batch_train[:, 0]] = batch_train[:, 2]
    final_movie_pop[batch_train[:, 1]] = batch_train[:, 3]

    if Reward_Base == 'ave_loss':
        user_total_losses[batch_train[:, 0], 1:last_num] = user_total_losses[batch_train[:, 0], 0:last_num-1]
        movie_total_losses[batch_train[:, 1], 1:last_num] = movie_total_losses[batch_train[:, 1], 0:last_num-1]
        user_total_losses[batch_train[:, 0], 0] = user_losses[batch_train[:, 0], 0]
        movie_total_losses[batch_train[:, 1], 0] = movie_losses[batch_train[:, 1], 0]

        user_count[batch_train[:, 0]] += 1
        movie_count[batch_train[:, 1]] += 1


if __name__ == "__main__":
    Output_Dim = 5 if Loss_Type == 'CrossEntropyLoss' else 1
    # Output_Dim = 5
    # if Loss_Type == 'CrossEntropyLoss':
    #     Output_Dim = 1

    # User_Num：用户数量
    # Movie_Num：电影数量
    train_features, test_features, train_target, test_target, genome_scores_dict, \
    train_feature_data, test_feature_data, Len_Train_Features, Len_Test_Features, \
    User_Num, Movie_Num, max_user_popularity, max_movie_popularity = load_data()
    train_feature_data, test_feature_data = train_feature_data[:Len_Train_Features], test_feature_data[:Len_Test_Features]

    Emb_Split = [0] + [sum(Emb_Size[0:i + 1]) for i in range(len(Emb_Size))]  # [0, 2, 18, 146]
    Setting_User_Popularity = [max_user_popularity, 32]
    Setting_Movie_Popularity = [max_movie_popularity, 32]
    Setting_User_Weight = [User_Num, len(Emb_Size)]
    Setting_Movie_Weight = [Movie_Num, len(Emb_Size)]

    if Train_Method == 'AutoML' and H_alpha > 0:
        Beta_Dis = nn.Embedding(num_embeddings=max(max_user_popularity, max_movie_popularity), embedding_dim=len(Emb_Size)).to(dtype=torch.float32)
        Beta_Dis.weight.data = torch.tensor(np.array([Beta(len(Emb_Size), popularity, Beta_Beta) for popularity in range(1, max(max_user_popularity, max_movie_popularity) + 1)]), dtype=torch.float32, requires_grad=False)
        Beta_Dis.weight.requires_grad = False
        Beta_Dis.to(device)
        criterion = nn.KLDivLoss(reduction='sum')

    user_policy = Policy(Setting_User_Popularity, Setting_User_Weight, Policy_Type)
    movie_policy = Policy(Setting_Movie_Popularity, Setting_Movie_Weight, Policy_Type)
    model = RS_MLP(Output_Dim, Dy_Emb_Num)
    user_policy.to(device)
    movie_policy.to(device)
    model.to(device)

    # 定义总参数量、可训练参数量及非可训练参数量变量
    Total_params = 0
    Trainable_params = 0
    NonTrainable_params = 0

    # # 遍历model.parameters()返回的全局参数列表
    # for param in movie_policy.parameters():
    #     mulValue = np.prod(param.size())  # 使用numpy prod接口计算参数数组所有元素之积
    #     Total_params += mulValue  # 总参数量
    #     if param.requires_grad:
    #         Trainable_params += mulValue  # 可训练参数量
    #     else:
    #         NonTrainable_params += mulValue  # 非可训练参数量
    #
    # print(f'Total params: {Total_params}')
    # print(f'Trainable params: {Trainable_params}')
    # print(f'Non-trainable params: {NonTrainable_params}')
    # print(model)


    if Model_Gpu:
        print('\n========================================================================================\n')
        print('Model_Gpu?:', next(model.parameters()).is_cuda, next(user_policy.parameters()).is_cuda, next(movie_policy.parameters()).is_cuda)
        print('Memory:    ', torch.cuda.memory_allocated(0) / 1024 ** 3, 'GB', torch.cuda.memory_cached(0) / 1024 ** 3, 'GB')
        print('\n========================================================================================\n')

    user_weights = np.zeros((User_Num,), dtype=np.int64)
    movie_weights = np.zeros((Movie_Num,), dtype=np.int64)
    final_user_pop = np.zeros((User_Num,), dtype=np.int64)
    final_movie_pop = np.zeros((User_Num,), dtype=np.int64)
    user_losses = np.ones((User_Num, 1), dtype=np.float32)
    movie_losses = np.ones((Movie_Num, 1), dtype=np.float32)

    if Reward_Base == 'ave_loss':
        user_total_losses = np.zeros((User_Num, last_num), dtype=np.float32)
        movie_total_losses = np.zeros((Movie_Num, last_num), dtype=np.float32)
        user_count = np.zeros((User_Num,), dtype=np.float32)
        movie_count = np.zeros((Movie_Num,), dtype=np.float32)

    t0 = time.time()

    optimizer_model = torch.optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=LR_model, weight_decay=0)
    optimizer_user  = torch.optim.Adam(filter(lambda p: p.requires_grad, user_policy.parameters()), lr=LR_darts, weight_decay=0)
    optimizer_movie = torch.optim.Adam(filter(lambda p: p.requires_grad, movie_policy.parameters()), lr=LR_darts, weight_decay=0)
    optimizer_darts = torch.optim.Adam(filter(lambda p: p.requires_grad, list(user_policy.parameters()) + list(movie_policy.parameters())), lr=LR_darts, weight_decay=0)
    optimizer_whole = torch.optim.Adam(filter(lambda p: p.requires_grad, list(model.parameters()) + list(user_policy.parameters()) + list(movie_policy.parameters())), lr=LR_model, weight_decay=0)
    losses = {'train': [], 'test': []}
    accuracies = {'train': [], 'test': []}
    train_sample_loss = list()
    train_sample_accuracy = list()
    user_dims_record = list()
    movie_dims_record = list()

    # 下面是训练部分
    print('\n******************************************Train******************************************\n')
    # trained_num_col = 1 # 存放已训练数量的列标号
    # loss_col = 2        # 存放当前损失度的列标号
    # accuracy_col = 3    # 存放当前训练精确度的列标号
    # row_num = 1         # 存放当前数据的行号，递增
    #
    # excel_name = "analyze/OEOPN_dim16.xlsx" # TODO Check
    # sheet_name = "ml_latest_bin" # TODO Check   ml_latest_multi   ml_latest_bin
    # excel = load_workbook(excel_name)
    # sheet = excel[sheet_name]

    for epoch_i in range(Epoch):
        index = 0
        while index < Len_Train_Features:
            update_controller(index, train_features, train_target)

            update_RS(index, train_features, Len_Train_Features, train_target, mode='train')

            if len(losses['train']) % 10 == 0:
                train_loss = sum(losses['train'][-10:]) / 10
                train_accuracy = sum([item[0] / item[1] for item in accuracies['train'][-10:]]) / 10
                print('Epoch = {:>3}  Batch = {:>4}/{:>4} ({:.3f}%)    train_loss = {:.3f}     train_accuracy = {:.3f}     total_time = {:.3f} min'.format(
                    epoch_i, index + Batch_Size, Len_Train_Features, 100 * (index + Batch_Size) / Len_Train_Features, train_loss,
                    train_accuracy, (time.time() - t0) / 60))
                # sheet.cell(row=row_num, column=trained_num_col).value = (epoch_i * Len_Train_Features) + index + Batch_Size
                # sheet.cell(row=row_num, column=loss_col).value = train_loss
                # sheet.cell(row=row_num, column=accuracy_col).value = train_accuracy
                # row_num = row_num + 1
            index += Batch_Size

    # save the excel
    # excel.close()
    # excel.save(excel_name)

    #############################test#############################
    print('\n******************************************Test******************************************\n')
    t0 = time.time()
    index = 0
    # Len_Test_Features = 20000
    times = 0
    test_accuracy_sum = 0
    while index < Len_Test_Features:
        update_controller(index, test_features, test_target)

        update_RS(index, test_features, Len_Test_Features, test_target, mode='test')

        # if len(losses['test']) % 10 == 0:
        #     print(
        #         'Test   Batch = {:>4}/{:>4} ({:.3f}%)     test_loss = {:.3f}     test_accuracy = {:.3f}     whole_time = {:.3f} min'.format(
        #             index + Batch_Size, Len_Test_Features, 100 * (index + Batch_Size) / Len_Test_Features,
        #             sum(losses['test'][-10:]) / 10,
        #             sum([item[0] / item[1] for item in accuracies['test'][-10:]]) / 10, (time.time() - t0) / 60))

        if len(losses['test']) % 10 == 0:
            test_accuracy = sum([item[0] / item[1] for item in accuracies['test'][-10:]]) / 10
            times += 1
            test_accuracy_sum += test_accuracy
            print(
                'Test   Batch = {:>4}/{:>4} ({:.3f}%)     test_loss = {:.3f}     test_accuracy = {:.3f}   avg_test_acc = {:.3f}   whole_time = {:.3f} min'.format(
                    index + Batch_Size, Len_Test_Features, 100 * (index + Batch_Size) / Len_Test_Features,
                    sum(losses['test'][-10:]) / 10, test_accuracy, test_accuracy_sum / times, (time.time() - t0) / 60))


        index += Batch_Size

    correct_num = sum([item[0] for item in accuracies['test']])
    test_num = sum([item[1] for item in accuracies['test']])

    print('Test Loss: {:.4f}'.format(sum(losses['test']) / test_num))

    print('Test Correct Num: {}'.format(correct_num))
    print('Test Num: {}'.format(test_num))

    print('Test Accuracy: {:.4f}'.format(correct_num / test_num))

    print("%s %s Batch_size=%d cross_emb_size=%d Epoch=%d PN_learning_rate=%.4f RS_learning_rate=%.3f Accuracy=%f Loss=%f" %
          (DATA_SET, Loss_Type, Batch_Size, cross_emb_size, Epoch, LR_darts, LR_model, (correct_num / test_num), (sum(losses['test']) / test_num)))

    # Save model
    save_model_name = './save_model/DyEmbNum{}_Policy_Type{}_LossType{}_Reward_Base{}_last{}_TestAcc{:.4f}'.format(
            Dy_Emb_Num, Policy_Type, Loss_Type, Reward_Base, last_num,
            correct_num / test_num)
    torch.save(model.state_dict(), save_model_name + '.pt')
    with open(save_model_name + '_weights.pkl', 'wb') as f:
        if Dy_Emb_Num == 1:
            pk.dump((final_user_pop, user_weights), f)
        elif Dy_Emb_Num == 2:
            pk.dump(((final_user_pop, user_weights), (final_movie_pop, movie_weights)), f)

    print('Model saved to ' + save_model_name + '.pt')
    print('Weights saved to ' + save_model_name + '_weights.pkl')

    feature_data = pd.concat([train_feature_data, test_feature_data])

    print("feature_data: ", feature_data.shape[0], feature_data.shape[1])

    feature_data['user_dims'] = pd.DataFrame(
        [[i] for i in user_dims_record])
    if Dy_Emb_Num == 2:
        feature_data['movie_dims'] = pd.DataFrame([[i] for i in movie_dims_record])
    feature_data['{}{}_loss_{}'.format(Train_Method[0], Policy_Type, Emb_Size)] = pd.DataFrame(
        [[i] for i in train_sample_loss])
    feature_data['{}{}_acc_{}'.format(Train_Method[0], Policy_Type, Emb_Size)] = pd.DataFrame(
        [[i] for i in train_sample_accuracy])

    print('\n****************************************************************************************\n')

    if Model_Gpu:
        print('\n++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n')
        print('Memory:    ', torch.cuda.memory_allocated(0) / 1024 ** 3, 'GB', torch.cuda.memory_cached(0) / 1024 ** 3, 'GB')
        # torch.cuda.empty_cache()
        print('Memory:    ', torch.cuda.memory_allocated(0) / 1024 ** 3, 'GB', torch.cuda.memory_cached(0) / 1024 ** 3, 'GB')
        print('\n++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n')

    Parameter_Name = 'DataSet{}_ValType{}_Policy{}_DyEmbNum{}_LossType{}_RewardBase{}'.format(
        DATA_SET,
        Val_Type if Train_Method == 'AutoML' else 'None',
        Policy_Type,
        Dy_Emb_Num,
        Loss_Type,
        Reward_Base)

    feature_data.to_csv('./results/feature_data_with_loss_{}.csv'.format(Parameter_Name), index=None)

    if Dy_Emb_Num == 1:
        result_user = []
        for i in range(1, 100):
            feature_data1 = feature_data[feature_data['user_frequency'] == i]
            result_user.append(list(feature_data1.mean(axis=0)) + [len(feature_data1)])
        Head = list(feature_data.columns) + ['count']
        pd.DataFrame(result_user).to_csv('./results/result_{}_user.csv'.format(Parameter_Name), index=None,
                                          header=Head)

    elif Dy_Emb_Num == 2:
        result_user, result_movie = [], []
        for i in range(1, 100):
            feature_data1 = feature_data[feature_data['user_frequency'] == i]
            result_user.append(list(feature_data1.mean(axis=0)) + [len(feature_data1)])
        Head = list(feature_data.columns) + ['count']
        pd.DataFrame(result_user).to_csv('./results/result_{}_user.csv'.format(Parameter_Name), index=None,
                                          header=Head)

        for i in range(1, 100):
            feature_data1 = feature_data[feature_data['movie_frequency'] == i]
            result_movie.append(list(feature_data1.mean(axis=0)) + [len(feature_data1)])
        Head = list(feature_data.columns) + ['count']
        pd.DataFrame(result_movie).to_csv('./results/result_{}_movie.csv'.format(Parameter_Name), index=None,
                                          header=Head)


    result = []
    for i in range(int(Train_Size/1000000)):
        feature_data1 = feature_data[i*1000000:(i+1)*1000000]
        result.append(list(feature_data1.mean(axis=0)) + [len(feature_data1)])

    Head = list(feature_data.columns) + ['count']
    pd.DataFrame(result).to_csv('./results/result_{}_trendency.csv'.format(Parameter_Name), index=None, header=Head)

    print('\n****************************************************************************************\n')
    print('os.getpid():   ', os.getpid())
    if torch.cuda.is_available():
        print('torch.cuda:    ', torch.cuda.is_available(), torch.cuda.current_device(), torch.cuda.device_count(), torch.cuda.get_device_name(0), torch.cuda.device(torch.cuda.current_device()))
    else:
        print('GPU is not available!!!')
    print('Train_Size:    ', Train_Size)
    print('Test_Size:     ', Test_Size)
    print('Emb_Size:      ', Emb_Size)
    print('Dy_Emb_Num:    ', Dy_Emb_Num)
    print('Loss_Type:     ', Loss_Type)
    print('Train_Method:  ', Train_Method)
    print('Policy_Type:   ', Types[Policy_Type])
    print('Val_Type:      ', Val_Type)
    print('Beta_Beta:     ', Beta_Beta)
    print('H_alpha:       ', H_alpha)
    print('LR_model:      ', LR_model)
    print('LR_darts:      ', LR_darts)
    print('\n****************************************************************************************\n')
    print('{} done'.format(Train_Method))
